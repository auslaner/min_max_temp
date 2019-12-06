"""
Determine minimum, maximum, and mean temperatures for each plant during
the time period that the plant was monitored.
"""
from datetime import datetime, timedelta
from statistics import mean

from openpyxl import load_workbook
from site_to_sheets import get_temp_sheet_name


def get_temp_file():
    """
    Loads file containing plant temperature data into openpyxl workbook object.
    :return: Openpyxl workbook object containing plant temperature data.
    """
    temp_file = load_workbook('/Users/u6000791/Box/Conservation/Rare Plants/Research Projects/Penstemon 2018-2019/'
                              'PENGRA_2019/DATA_PENGRA_2019/Measured variables_PENGRA.xlsx')
    return temp_file


def get_output_file():
    output_file = load_workbook('/Users/u6000791/Desktop/PENGRA_SEM.xlsx')
    return output_file


def get_next_plant(plants_ws):
    """
    Return row containing next plant info.
    :param plants_ws: Openpyxl worksheet containing monitored plant info.
    :return: Openpyxl row.
    """
    for row in plants_ws.iter_rows(min_row=3, max_row=243, max_col=9):
        yield row


def get_site(plant_row):
    """
    Get the site the plant is from
    :param plant_row: Openpyxl describing a monitored plant.
    :return: The string abbreviation of the site name where the plant is from.
    """
    # Return site info which is stored in the first column of the row.
    return plant_row[0].value


def get_start_time(plant_row, monitoring_date):
    """
    Get the monitoring start time for the given plant
    :param monitoring_date: A datetime object containing the day, month, and
    year that monitoring took place.
    :param plant_row: Openpyxl describing a monitored plant.
    :return: Datetime object representing when plant monitoring began.
    """
    # Start time data is in the 5th column
    return datetime(day=monitoring_date.day, month=monitoring_date.month,
                    year=monitoring_date.year, hour=plant_row[4].value.hour,
                    minute=plant_row[4].value.minute)


def get_monitoring_date(plant_row):
    """
    Get the monitoring date for the given plant
    :param plant_row: Openpyxl describing a monitored plant.
    :return: Datetime object representing the date of monitoring.
    """
    # Date value is in the 3rd column
    return plant_row[2].value


def get_end_time(plant_row, monitoring_date):
    """
    Get the monitoring end time for the given plant
    :param monitoring_date: A datetime object containing the day, month, and
    year that monitoring took place.
    :param plant_row: Openpyxl describing a monitored plant.
    :return: Datetime object representing when plant monitoring began.
    """
    # End time data is in the 6th column
    return datetime(day=monitoring_date.day, month=monitoring_date.month,
                    year=monitoring_date.year, hour=plant_row[5].value.hour,
                    minute=plant_row[5].value.minute)


def filter_temps_by_monitoring(temp_range, start_time, end_time):
    """
    Return a list of temperatures that were recorded only within the
    monitored period. Note that the considered monitored period is
    expanded to 15 minutes before and after the given start and end
    times so as to include recorded temperatures close to when
    monitoring began and ended.
    :param temp_range: Cell range containing datetimes and recorded
    temperatures in degrees Centigrade.
    :param start_time: Datetime when monitoring began.
    :param end_time: Datetime when monitoring ended.
    :return: List of temperatures during monitoring period.
    """
    monitored_temps = []
    for row in temp_range:
        try:
            # Check if the datetime in the row's first column is in between
            # the start and end monitoring time
            temp_date = datetime.strptime(row[0].value, '%Y-%m-%d %H:%M:%S')
        except TypeError:
            # We're parsing a sheet that has fewer than 1440 rows of temp data
            break
        if start_time - timedelta(minutes=15) <= temp_date <= end_time + timedelta(minutes=15):
            monitored_temps.append(row[1].value)

    return monitored_temps


def compute_from_temp_range(temp_range):
    """
    Return the minimum, maximum, and mean temperatures from the list
    of temperatures.
    :param temp_range: List of temperatures in degrees Centigrade.
    :return: Minimum, maximum, and mean temperatures in degrees Centigrade.
    """
    minimum_temp = min(temp_range)
    maximum_temp = max(temp_range)
    mean_temp = round(mean(temp_range), 1)
    return minimum_temp, maximum_temp, mean_temp


def save_temp_stats(plant_row, temp_min, temp_max, temp_mean):
    """
    Save the temperatures stats back to the plant workbook
    :param plant_row: Openpyxl describing a monitored plant.
    :param temp_min: Minimum temperature in degrees Centigrade.
    :param temp_max: Maximum temperature in degrees Centigrade.
    :param temp_mean: Mean temperature in degrees Centigrade.
    :return: None.
    """
    # Save min temp to 7th column
    plant_row[6].value = temp_min

    # Save max temp to 8th column
    plant_row[7].value = temp_max

    # Save mean temp to 9th column
    plant_row[8].value = temp_mean


def main():
    # Get the workbook from our file with the plant numbers and monitoring times
    output_wb = get_output_file()

    # Get the file containing the temperature data
    temp_wb = get_temp_file()

    # Get the sheet with our plants from the output workbook
    plants_ws = output_wb['Temp']

    # Get a row from the plant worksheet
    for plant_row in get_next_plant(plants_ws):
        # Get the date of monitoring
        monitoring_date = get_monitoring_date(plant_row)

        # Get the monitoring start time for the plant
        start_time = get_start_time(plant_row, monitoring_date)

        # Get the monitoring end time for the plant
        end_time = get_end_time(plant_row, monitoring_date)

        # Determine which site the plant described in the row is from
        site = get_site(plant_row)

        # Get the name of the sheet containing temp data for this site
        temp_sheet_name = get_temp_sheet_name(site)

        # Get the actual worksheet of temp data now that we have the name
        temp_sheet = temp_wb[temp_sheet_name]

        # Get all of the datetimes and temperatures from the sheet
        temp_range = temp_sheet['B27:C1440']  # Note that 1440 is chosen because it is the highest row count among sites

        # Filter temp range by monitoring start and end times
        filtered_temp_range = filter_temps_by_monitoring(temp_range, start_time, end_time)

        # Compute the min, max, and mean from our filtered list of temps
        temp_min, temp_max, temp_mean = compute_from_temp_range(filtered_temp_range)

        # Save temp vales to output worksheet
        save_temp_stats(plant_row, temp_min, temp_max, temp_mean)

    # Save the worksheet with the new temperature stats added
    output_wb.save('PENGRA_monitored_temps.xlsx')


if __name__ == "__main__":
    main()
